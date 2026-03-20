import re
import os
from docx import Document

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


def extract_field(text, label):
    """
    Enhanced field extraction supporting multiple delimiters:
    - 'Label- Value'
    - 'Label: Value'
    - 'Label = Value'
    - 'Label: Value' (with colon)
    Works with multi-word values like full names.
    """
    escaped = re.escape(label)
    # Try multiple delimiter patterns: dash, colon, equals, and space-separated
    patterns = [
        rf"{escaped}\s*[-–—]\s*(.+?)(?:\r?\n|$)",      # dash variants
        rf"{escaped}\s*:\s*(.+?)(?:\r?\n|$)",           # colon
        rf"{escaped}\s*=\s*(.+?)(?:\r?\n|$)",           # equals
        rf"{escaped}\s+([^\n:=\-]+?)(?:\r?\n|$)",       # space-separated
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return "—"


def extract_from_docx(filepath, fields):
    """
    Reads a .docx file and extracts the requested fields.
    fields: list of label strings e.g. ["Name", "Serial No."]
    Returns a dict with 'file' key plus one key per field.
    """
    try:
        doc = Document(filepath)
        # Join all paragraphs with newlines so regex can find fields anywhere
        text = "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        # If file can't be read, return empty result
        result = {"file": os.path.basename(filepath), "error": str(e), "format": "docx"}
        for f in fields:
            result[f] = "—"
        return result

    result = {"file": os.path.basename(filepath), "format": "docx"}

    found_count = 0
    for field in fields:
        value = extract_field(text, field)
        result[field] = value
        if value != "—":
            found_count += 1

    # Confidence = how many fields were actually found
    result["confidence"] = round(found_count / len(fields), 2) if fields else 0.0

    return result


def extract_from_pdf(filepath, fields):
    """
    Extracts text from PDF files using PyPDF2 or pdfplumber.
    Falls back to text extraction if library available.
    """
    text = ""

    try:
        if PDFPLUMBER_AVAILABLE:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
        elif PDF_AVAILABLE:
            import PyPDF2
            with open(filepath, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
        else:
            result = {"file": os.path.basename(filepath), "format": "pdf", "error": "PDF library not available"}
            for f in fields:
                result[f] = "—"
            return result

    except Exception as e:
        result = {"file": os.path.basename(filepath), "format": "pdf", "error": str(e)}
        for f in fields:
            result[f] = "—"
        return result

    result = {"file": os.path.basename(filepath), "format": "pdf"}

    found_count = 0
    for field in fields:
        value = extract_field(text, field)
        result[field] = value
        if value != "—":
            found_count += 1

    result["confidence"] = round(found_count / len(fields), 2) if fields else 0.0
    return result


def extract_from_txt(filepath, fields):
    """
    Extracts data from plain text files.
    """
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            text = file.read()
    except Exception as e:
        result = {"file": os.path.basename(filepath), "format": "txt", "error": str(e)}
        for f in fields:
            result[f] = "—"
        return result

    result = {"file": os.path.basename(filepath), "format": "txt"}

    found_count = 0
    for field in fields:
        value = extract_field(text, field)
        result[field] = value
        if value != "—":
            found_count += 1

    result["confidence"] = round(found_count / len(fields), 2) if fields else 0.0
    return result


def extract_from_xlsx(filepath, fields):
    """
    Extracts data from Excel files (.xlsx, .xls).
    """
    if not XLSX_AVAILABLE:
        result = {"file": os.path.basename(filepath), "format": "xlsx", "error": "openpyxl not available"}
        for f in fields:
            result[f] = "—"
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
    except Exception as e:
        result = {"file": os.path.basename(filepath), "format": "xlsx", "error": str(e)}
        for f in fields:
            result[f] = "—"
        return result

    result = {"file": os.path.basename(filepath), "format": "xlsx"}

    found_count = 0
    for field in fields:
        value = extract_field(text, field)
        result[field] = value
        if value != "—":
            found_count += 1

    result["confidence"] = round(found_count / len(fields), 2) if fields else 0.0
    return result


def process_folder(folder_path, fields):
    """
    Processes all supported document files in a folder.
    Supports: .docx, .pdf, .xlsx, .txt
    folder_path: string path to the folder
    fields: list of field label strings to extract
    Returns list of dicts, one per file.
    """
    results = []

    if not os.path.exists(folder_path):
        return []

    # Supported file extensions with their handler functions
    handlers = {
        ".docx": extract_from_docx,
        ".pdf": extract_from_pdf,
        ".xlsx": extract_from_xlsx,
        ".xls": extract_from_xlsx,
        ".txt": extract_from_txt,
    }

    files = sorted([
        f for f in os.listdir(folder_path)
        if any(f.lower().endswith(ext) for ext in handlers.keys())
    ])

    for fname in files:
        filepath = os.path.join(folder_path, fname)
        # Get the file extension and find the appropriate handler
        ext = os.path.splitext(fname)[1].lower()

        if ext in handlers:
            try:
                result = handlers[ext](filepath, fields)
                results.append(result)
            except Exception as e:
                result = {"file": fname, "format": ext[1:], "error": str(e), "confidence": 0.0}
                for f in fields:
                    result[f] = "—"
                results.append(result)

    return results


# ── Quick test ────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    folder = "demo_files"
    fields = ["Name", "Serial No."]

    print(f"Scanning: {folder}")
    print(f"Fields:   {fields}\n")

    results = process_folder(folder, fields)

    for r in results:
        print(f"{r['file']:<25}", end="")
        for f in fields:
            print(f"  {f}: {r[f]:<30}", end="")
        print(f"  confidence: {r['confidence']}")

    found = sum(1 for r in results if r["confidence"] == 1.0)
    print(f"\n{found}/{len(results)} files fully extracted") 